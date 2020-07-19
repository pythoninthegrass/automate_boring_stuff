#!/usr/bin/env python3

import openpyxl
import os

# `cd` to directory with *.xlsx
cwd = os.getcwd() + '/resources'
os.chdir(cwd)

# load workbook with values-only (not formulas)
wb = openpyxl.load_workbook('time_sheet.xlsx', data_only=True)
type(wb)                                                        # <class 'openpyxl.workbook.workbook.Workbook'>
wb.get_sheet_names()                                            # ['Sheet1']
sheet = wb['Sheet1']                                            # Previously `wb.get_sheet_by_name('Sheet1')`

# sheet object + value
cell = sheet['A1']
cell.value                                                      # 'Massachusetts Institute of Technology'

str(sheet['A1'].value)                                          # If `data_only=True` isn't passed, avoids interpolation of formulas (e.g., `datetime.time(9, 0)`)
sheet['A4'].value                                               # "Classified Employee's Time and Attendance Record"
str(sheet['C19'].value)                                         # '09:00:00'
sheet.cell(row=16, column=2)                                    # <Cell 'Sheet1'.B16>

# iterate through col B, rows 28 - 34
for i in range(28, 35):
    print(i, sheet.cell(row=i, column=2).value)                 # '28 2013-12-14 00:00:00 ...'