#!/usr/bin/env python3

import openpyxl
import os

# `cd` to directory with *.xlsx
cwd = os.getcwd() + '/resources'
os.chdir(cwd)

# create workbook + sheet (in-memory)
wb = openpyxl.Workbook()
wb.sheetnames
sheet = wb['Sheet']

# assign values to cells
sheet['A1'].value == None                           # True
sheet['A1'] = 42
sheet['A1'].value                                   # 42
sheet['A2'] = 'hello'

# create new sheet
sheet2 = wb.create_sheet(title='Sheet2')            # creat_sheet(index=0, title='...') moves new sheet to the top (far-left)
sheet2.title = 'Formerly known as Sheet2'
wb.sheetnames                                       # ['Sheet', 'Formerly known as Sheet2']

# save to xlsx
wb.save('example.xlsx')